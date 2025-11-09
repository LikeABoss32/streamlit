# app_final_v5.py
import streamlit as st
import pandas as pd, numpy as np
import plotly.express as px, plotly.graph_objects as go
from io import BytesIO
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.preprocessing import LabelEncoder
from sklearn.inspection import permutation_importance
import base64, time
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader

st.set_page_config(layout="wide", page_title="Disease Dashboard (v5)", initial_sidebar_state="collapsed")

# --- Styling: pink -> purple gradient like PDF screenshots ---
st.markdown("""
    <style>
    .stApp { background: linear-gradient(180deg, #ff9a9e 0%, #8e44ad 100%); }
    .header { background: rgba(255,255,255,0.06); padding:18px; border-radius:8px; margin-bottom:10px; }
    .title { font-size:28px; color:white; font-weight:700; text-align:center; }
    .card { background:white; border-radius:12px; padding:14px; box-shadow:0 8px 30px rgba(0,0,0,0.12); }
    .kpi { text-align:center; padding:12px; border-radius:8px; color:white; font-weight:700; }
    .btn { background: linear-gradient(90deg,#ff6aa3,#8e44ad); color:white; padding:8px 12px; border-radius:8px; }
    </style>
""", unsafe_allow_html=True)

@st.cache_data
def load_data(path=r"C:\Users\Karan\Downloads\dashboard\disease_data_final_v5.csv"):
    return pd.read_csv(path)

df = load_data()

# Header
st.markdown('<div class="header"><div class="title">Disease Analysis & Monitoring Dashboard </div></div>', unsafe_allow_html=True)

# Filters
with st.form("filters"):
    c1,c2,c3,c4,c5 = st.columns([1.4,1.4,1,1,0.8])
    with c1:
        state = st.selectbox("State/UT", ["All"] + sorted(df['state'].unique()))
    with c2:
        city = st.selectbox("City", ["All"] + sorted(df['city'].unique()))
    with c3:
        year = st.selectbox("Year", ["All"] + sorted(df['year'].unique()))
    with c4:
        disease = st.selectbox("Disease", ["All"] + sorted(df['disease'].unique()))
    with c5:
        submit = st.form_submit_button("UPDATE DASHBOARD")
if submit:
    st.rerun()


# Apply filters
d = df.copy()
if state != "All":
    d = d[d['state']==state]
if city != "All":
    d = d[d['city']==city]
if year != "All":
    d = d[d['year']==int(year)]
if disease != "All":
    d = d[d['disease']==disease]

# KPI Tiles with fade-in (simulate by small time delays for presentation)
k1,k2,k3,k4 = st.columns([1,1,1,1])
k1.write("")
k2.write("")
k3.write("")
k4.write("")
k1.markdown(f'<div class="card"><div class="kpi" style="background:linear-gradient(90deg,#ff6aa3,#ffb3e6);">Total Cases<br><span style="font-size:22px;font-weight:800">{d["cases"].sum():,}</span></div></div>', unsafe_allow_html=True)
k2.markdown(f'<div class="card"><div class="kpi" style="background:linear-gradient(90deg,#8e44ad,#b28cff);">Total Deaths<br><span style="font-size:22px;font-weight:800">{d["deaths"].sum():,}</span></div></div>', unsafe_allow_html=True)
k3.markdown(f'<div class="card"><div class="kpi" style="background:linear-gradient(90deg,#ff9a9e,#ff6aa3);">Affected States<br><span style="font-size:22px;font-weight:800">{d["state"].nunique()}</span></div></div>', unsafe_allow_html=True)
k4.markdown(f'<div class="card"><div class="kpi" style="background:linear-gradient(90deg,#6a11cb,#ff6a00);">Avg Temp (K)<br><span style="font-size:22px;font-weight:800">{round(d["temperature"].mean(),2) if not d.empty else 0}</span></div></div>', unsafe_allow_html=True)
time.sleep(0.15)

st.markdown("<br/>", unsafe_allow_html=True)

# Tabs
tabs = st.tabs(["Overview","Geographic Analysis","Trend Analysis","ML Prediction","Reports & Downloads"])

# --- Overview tab ---
with tabs[0]:
    st.markdown('<div class="card"><h3>Cases Timeline</h3></div>', unsafe_allow_html=True)
    timeline = d.groupby('year', as_index=False)['cases'].sum().sort_values('year')
    if timeline.empty:
        st.info("No data")
    else:
        fig = px.bar(timeline, x='year', y='cases', color_discrete_sequence=['#ff6aa3'])
        fig.update_layout(height=420, transition={'duration':700})
        st.plotly_chart(fig, use_container_width=True)
    st.markdown('<div class="card" style="margin-top:12px"><div style="display:flex;gap:18px"><div style="flex:1">', unsafe_allow_html=True)
        # --- Dynamic Seasonal Analysis and Pie Chart (Fully Reactive) ---
if not d.empty and "month" in d.columns:
    # Define Seasons Function
    def month_to_season(m):
        if m in [6, 7, 8, 9]:
            return "Monsoon"
        elif m in [10, 11, 12, 1]:
            return "Post-Monsoon"
        elif m in [2, 3, 4]:
            return "Summer"
        else:
            return "Winter"

    # Ensure we create a fresh copy and apply the season label
    d = d.copy()
    d["season"] = d["month"].apply(month_to_season)

    # Define consistent season order
    season_order = ["Monsoon", "Post-Monsoon", "Summer", "Winter"]

    # Group safely and reindex to include all seasons
    seasonal = (
        d.groupby("season", as_index=False)["cases"]
        .mean()
        .set_index("season")
        .reindex(season_order, fill_value=0)
        .reset_index()
    )

    # --- Premium Seasonal Radar Chart (Enhanced Visibility) ---
    fig_rad = go.Figure()

    season_colors = {
        "Monsoon": "rgba(255, 20, 147, 0.9)",      # Deep Pink
        "Post-Monsoon": "rgba(138, 43, 226, 0.9)", # Blue Violet
        "Summer": "rgba(0, 191, 255, 0.9)",        # Sky Blue
        "Winter": "rgba(255, 140, 0, 0.9)"         # Orange
    }

    theta_values = season_order
    r_values = seasonal["cases"].tolist()
    colors = [season_colors[s] for s in theta_values]

    fig_rad.add_trace(go.Scatterpolar(
        r=r_values,
        theta=theta_values,
        fill="toself",
        name="Average Cases",
        line_color="rgba(255,255,255,1)",
        fillcolor="rgba(138,43,226,0.5)",  # Bright purple-pink
        marker=dict(size=9, color=colors, line=dict(color="white", width=1.2)),
        hovertemplate="<b>%{theta}</b><br>Avg Cases: %{r:.0f}<extra></extra>",
    ))

    fig_rad.update_layout(
        polar=dict(
            bgcolor="rgba(30,0,60,0.9)",
            radialaxis=dict(
                visible=True,
                showgrid=True,
                gridcolor="rgba(255,255,255,0.15)",
                linecolor="rgba(255,255,255,0.2)",
                tickfont=dict(color="white", size=13),
            ),
            angularaxis=dict(
                tickfont=dict(color="white", size=14, family="Arial Black"),
                gridcolor="rgba(255,255,255,0.15)",
            ),
        ),
        showlegend=False,
        height=420,
        paper_bgcolor="rgba(40,0,80,1)",
        font=dict(color="white", size=13),
        margin=dict(t=20, b=20, l=20, r=20),
        title=dict(
            text="Seasonal Disease Trend",
            x=0.5,
            font=dict(size=16, color="white", family="Arial Black"),
        ),
    )

    st.plotly_chart(fig_rad, use_container_width=True)
    st.markdown(
        "</div><div style='width:18px'></div><div style='flex:1'>",
        unsafe_allow_html=True,
    )

    # --- High-Contrast Disease Proportion Pie Chart ---
    pie = d.groupby("disease", as_index=False)["cases"].sum()
    if not pie.empty:
        vibrant_colors = [
            "#FF007F", "#00FFFF", "#FF8C00", "#9932CC",
            "#FF1493", "#00FF7F", "#FFD700"
        ]
        fig_p = px.pie(
            pie,
            names="disease",
            values="cases",
            color_discrete_sequence=vibrant_colors,
        )
        fig_p.update_traces(
            textinfo="label+percent",
            textfont_size=13,
            pull=[0.05] * len(pie),
        )
        fig_p.update_layout(
            height=400,
            paper_bgcolor="rgba(40,0,80,1)",
            font=dict(color="white", size=13),
            title=dict(
                text="Disease-Wise Proportion",
                x=0.5,
                font=dict(size=16, color="white", family="Arial Black"),
            ),
        )
        st.plotly_chart(fig_p, use_container_width=True)
else:
    st.warning("No data available for selected filters to display seasonal analysis.")

    # Top 10 cities horizontal bar
    st.markdown('<div class="card" style="margin-top:12px"><h4>Top 10 Cities by Cases</h4>', unsafe_allow_html=True)
    top_cities = d.groupby('city', as_index=False)['cases'].sum().sort_values('cases', ascending=False).head(10)
    if not top_cities.empty:
        fig_tc = px.bar(top_cities, x='cases', y='city', orientation='h', color='cases', color_continuous_scale=['#ff9a9e','#8e44ad'])
        fig_tc.update_layout(height=360)
        st.plotly_chart(fig_tc, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

# --- Geographic Analysis ---
with tabs[1]:
    st.markdown('<div class="card"><h4>Geographic Distribution (Longitude vs Latitude)</h4></div>', unsafe_allow_html=True)
    geo = d.groupby(['city','latitude','longitude'], as_index=False)['cases'].sum()
    if geo.empty:
        st.info("No data")
    else:
        fig_map = px.scatter_mapbox(geo, lat='latitude', lon='longitude', size='cases', hover_name='city', color='cases', color_continuous_scale='RdPu', zoom=4, height=430)
        fig_map.update_layout(mapbox_style='open-street-map', margin=dict(l=10,r=10,t=10,b=10))
        st.plotly_chart(fig_map, use_container_width=True)
    # 3D correlation
    st.markdown('<div class="card" style="margin-top:12px"><h4>3D Climate-Disease Correlation</h4></div>', unsafe_allow_html=True)
    agg3 = d.groupby('city', as_index=False).agg({'cases':'sum','temperature':'mean','precipitation':'mean'})
    if not agg3.empty:
        fig3d = px.scatter_3d(agg3, x='temperature', y='precipitation', z='cases', color='cases', hover_name='city', height=420, color_continuous_scale='RdPu')
        st.plotly_chart(fig3d, use_container_width=True)
    # bubble correlation
    st.markdown('<div class="card" style="margin-top:12px"><h4>Bubble: Cases vs Temp vs Precipitation</h4></div>', unsafe_allow_html=True)
    bubble = d.groupby('city', as_index=False).agg({'cases':'sum','temperature':'mean','precipitation':'mean'})
    if not bubble.empty:
        fig_b = px.scatter(bubble, x='temperature', y='precipitation', size='cases', color='cases', hover_name='city', color_continuous_scale='RdPu', height=420)
        st.plotly_chart(fig_b, use_container_width=True)

# --- Trend Analysis ---
with tabs[2]:
    st.markdown('<div class="card"><h4>Deaths Timeline and Climate Factors</h4></div>', unsafe_allow_html=True)
    dt = d.groupby('year', as_index=False)['deaths'].sum().sort_values('year')
    figd = px.line(dt, x='year', y='deaths', markers=True, color_discrete_sequence=['#b28cff'])
    figd.update_layout(height=300)
    st.plotly_chart(figd, use_container_width=True)
    climate = d.groupby('month', as_index=False)[['temperature','precipitation']].mean().sort_values('month')
    if not climate.empty:
        figc = go.Figure()
        figc.add_trace(go.Scatter(x=climate['month'], y=climate['temperature'], mode='lines+markers', name='Temperature (K)', line=dict(color='#ff6aa3')))
        figc.add_trace(go.Bar(x=climate['month'], y=climate['precipitation'], name='Precipitation (mm)', opacity=0.6, yaxis='y2', marker_color='#8e44ad'))
        figc.update_layout(yaxis2=dict(overlaying='y', side='right'), height=380)
        st.plotly_chart(figc, use_container_width=True)
    # State-wise heatmap (pivot)
    st.markdown('<div class="card" style="margin-top:12px"><h4>State-wise Heatmap (Total Cases)</h4></div>', unsafe_allow_html=True)
    state_heat = d.groupby(['state','year'], as_index=False)['cases'].sum().pivot(index='state', columns='year', values='cases').fillna(0)
    if not state_heat.empty:
        fig_h = px.imshow(state_heat.values, x=state_heat.columns, y=state_heat.index, color_continuous_scale='RdPu', aspect='auto')
        fig_h.update_layout(height=420)
        st.plotly_chart(fig_h, use_container_width=True)
    # Yearly stacked bar by disease
    st.markdown('<div class="card" style="margin-top:12px"><h4>Yearly Stacked: Disease vs Year</h4></div>', unsafe_allow_html=True)
    ystack = d.groupby(['year','disease'], as_index=False)['cases'].sum()
    if not ystack.empty:
        fig_yst = px.bar(ystack, x='year', y='cases', color='disease', height=420)
        st.plotly_chart(fig_yst, use_container_width=True)

# --- ML Prediction (enhanced visuals) ---
with tabs[3]:
    st.markdown('<div class="card"><h4>Enhanced ML Prediction (Top-5 probabilities, feature impact, risk gauge)</h4></div>', unsafe_allow_html=True)
    # Prepare training data
    data_ml = df.copy()
    data_ml = data_ml.sort_values(['city','disease','year','month','week'])
    data_ml['prev_week_cases'] = data_ml.groupby(['city','disease'])['cases'].shift(1).fillna(0)
    features = ['month','temperature','precipitation','lai','prev_week_cases']
    le = LabelEncoder()
    data_ml['disease_code'] = le.fit_transform(data_ml['disease'])
    # sample to speed up training
    train = data_ml.dropna().sample(n=min(6000, len(data_ml)), random_state=42)
    X = train[features]
    y = train['disease_code']
    clf = GradientBoostingClassifier(n_estimators=120, random_state=42)
    try:
        clf.fit(X, y)
    except Exception as e:
        st.error("Model training failed: " + str(e))
    # compute permutation importance (approximate) for visualization
    try:
        imp = permutation_importance(clf, X, y, n_repeats=6, random_state=42, n_jobs=1)
        importances = pd.Series(imp.importances_mean, index=features).sort_values(ascending=False)
    except Exception:
        importances = pd.Series(clf.feature_importances_, index=features).sort_values(ascending=False)
    # Prediction inputs
    with st.form("predict_form"):
        p1,p2,p3,p4 = st.columns([1.2,1,1,1])
        with p1:
            in_city = st.selectbox("City for prediction", ["All"] + sorted(df['city'].unique()), index=1)
        with p2:
            in_month = st.selectbox("Month", list(range(1,13)), index=6)
        with p3:
            in_temp = st.number_input("Temperature (K)", value=303.0)
        with p4:
            in_prec = st.number_input("Precipitation (mm)", value=2.5)
        in_lai = st.number_input("LAI", value=3.0)
        submit_pred = st.form_submit_button("Predict Top Diseases")
    if submit_pred:
        if in_city != "All":
            prev = df[(df['city']==in_city)].sort_values(['year','month','week']).tail(7)['cases'].mean()
        else:
            prev = df['cases'].tail(7).mean()
        Xpred = pd.DataFrame([{'month':in_month,'temperature':in_temp,'precipitation':in_prec,'lai':in_lai,'prev_week_cases':prev}])
        probs = clf.predict_proba(Xpred)[0]
        top_idx = np.argsort(probs)[::-1][:5]
        diseases_list = le.inverse_transform(top_idx)
        top_probs = probs[top_idx]
        # show bar chart of top 5 probabilities
        prob_df = pd.DataFrame({'disease': diseases_list, 'prob': top_probs})
        fig_probs = px.bar(prob_df, x='disease', y='prob', color='prob', color_continuous_scale='RdPu', range_y=[0,1], height=320)
        st.plotly_chart(fig_probs, use_container_width=True)
        # feature impact radar using importances
        fig_feat = go.Figure()
        vals = importances.reindex(features).fillna(0).values
        fig_feat.add_trace(go.Scatterpolar(r=vals, theta=features, fill='toself', name='Feature importance', line_color='#ff6aa3'))
        fig_feat.update_layout(polar=dict(radialaxis=dict(visible=True)), height=360)
        st.plotly_chart(fig_feat, use_container_width=True)
        # risk gauge using indicator
        risk_score = float(top_probs[0])
        gauge = go.Figure(go.Indicator(mode="gauge+number", value=risk_score*100,
                                      gauge={'axis': {'range': [0,100]}, 'bar': {'color':'#ff6aa3'},
                                             'steps':[{'range':[0,35],'color':'#90ee90'},{'range':[35,60],'color':'#ffd700'},{'range':[60,100],'color':'#ff6aa3'}]}))
        gauge.update_layout(height=240)
        st.plotly_chart(gauge, use_container_width=True)
        # explanation text (simple)
        if risk_score > 0.6:
            risk_cat = "High"
        elif risk_score > 0.35:
            risk_cat = "Medium"
        else:
            risk_cat = "Low"
        st.markdown(f"**Top prediction:** {diseases_list[0]} ({top_probs[0]:.1%}), **Risk:** {risk_cat}")
        # show full probability table
        full_probs = pd.DataFrame({'disease': le.inverse_transform(np.arange(len(le.classes_))), 'prob': clf.predict_proba(Xpred)[0]}).sort_values('prob', ascending=False)
        st.dataframe(full_probs.head(12))
    st.markdown('<div style="height:12px"></div>', unsafe_allow_html=True)

# --- Reports & Downloads ---
with tabs[4]:
    st.markdown('<div class="card"><h4>Reports & Downloads</h4></div>', unsafe_allow_html=True)
    st.download_button("Download Filtered CSV", d.to_csv(index=False).encode('utf-8'), "filtered_data_v5.csv", "text/csv")
    # Excel summary generator
    def make_excel_summary(filtered_df):
        wb = Workbook()
        # Sheet1: Filtered Data
        ws1 = wb.active
        ws1.title = "Filtered_Data"
        for r in dataframe_to_rows(filtered_df, index=False, header=True):
            ws1.append(r)
        # style header
        for cell in ws1[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF6AA3", end_color="FF6AA3", fill_type="solid")
        # Sheet2: Year Summary
        ws2 = wb.create_sheet("Year_Summary")
        ys = filtered_df.groupby('year', as_index=False)['cases','deaths'].sum().sort_values('year')
        for r in dataframe_to_rows(ys, index=False, header=True):
            ws2.append(r)
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        # Sheet3: Disease Summary
        ws3 = wb.create_sheet("Disease_Summary")
        ds = filtered_df.groupby('disease', as_index=False)['cases','deaths'].sum().sort_values('cases', ascending=False)
        for r in dataframe_to_rows(ds, index=False, header=True):
            ws3.append(r)
        for cell in ws3[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        # Sheet4: Seasonal Average
        ws4 = wb.create_sheet("Seasonal_Avg")
        def month_to_season(m):
            if m in [6,7,8,9]: return "Monsoon"
            if m in [10,11,12,1]: return "Post-Monsoon"
            if m in [2,3,4]: return "Summer"
            return "Winter"
        filtered_df['season'] = filtered_df['month'].apply(month_to_season)
        sa = filtered_df.groupby('season', as_index=False)['cases','deaths'].mean()
        for r in dataframe_to_rows(sa, index=False, header=True):
            ws4.append(r)
        for cell in ws4[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        # Save to bytes
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio.getvalue()
    if st.button("Download Excel Summary (multi-sheet)"):
        excel_bytes = make_excel_summary(d.copy())
        b64 = base64.b64encode(excel_bytes).decode()
        href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="disease_summary_v5.xlsx">Click here to download Excel summary</a>'
        st.markdown(href, unsafe_allow_html=True)
    # PDF report (simple)
    def make_pdf(filtered):
        buf = BytesIO()
        c = canvas.Canvas(buf, pagesize=A4)
        w,h = A4
        c.setFont("Helvetica-Bold", 16)
        c.drawString(40, h-60, "Disease Analysis Report (v5)")
        c.setFont("Helvetica", 11)
        c.drawString(40, h-90, f"Total cases: {filtered['cases'].sum():,}")
        c.drawString(40, h-110, f"Total deaths: {filtered['deaths'].sum():,}")
        try:
            import plotly.io as pio
            t = filtered.groupby('year', as_index=False)['cases'].sum()
            fig = px.bar(t, x='year', y='cases', title='Cases timeline', color_discrete_sequence=['#ff6aa3'])
            img = pio.to_image(fig, format='png', width=640, height=260)
            img_reader = ImageReader(BytesIO(img))
            c.drawImage(img_reader, 40, h-370, width=520, height=200)
        except Exception:
            c.drawString(40, h-130, "Charts not embedded (kaleido missing).")
        c.showPage()
        c.save()
        buf.seek(0)
        return buf.read()
    if st.button("Generate PDF Report"):
        pdf = make_pdf(d.copy())
        b64 = base64.b64encode(pdf).decode()
        href = f'<a href="data:application/pdf;base64,{b64}" download="disease_report_v5.pdf">Download PDF report</a>'
        st.markdown(href, unsafe_allow_html=True)
    st.markdown('<div style="height:12px"></div>', unsafe_allow_html=True)

# Raw data preview (small)
st.markdown('<div class="card"><h4>Data Preview (filtered)</h4></div>', unsafe_allow_html=True)
st.dataframe(d.head(200))